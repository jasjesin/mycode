import openpyxl

inv_file = openpyxl.load_workbook("inventory1.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
mkt_cap_per_supplier = {}
low_inventory = {}
product_amount = {}

for row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(row, 4).value
    inventory = int (product_list.cell(row, 2).value)
    price = int(product_list.cell(row, 3).value)
    product = int(product_list.cell(row, 1).value)
    product_total = product_list.cell(row, 5)

    # calculate total number of products offered per supplier
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    # total value of inventory held per supplier
    if supplier_name in mkt_cap_per_supplier:
        current_total = mkt_cap_per_supplier.get(supplier_name)
        mkt_cap_per_supplier[supplier_name] = current_total + inventory * price
    else:
        mkt_cap_per_supplier[supplier_name] = int(inventory * price)

    # look for low inventory under 100 to re-order
    if inventory < 100:
        low_inventory[product] = int(inventory)

    # Total amount spent on each product
    product_total.value = inventory * price
    inv_file.save("inventory2.xlsx")

print("\n==> Total number of products per supplier:", products_per_supplier)
print("\n==> Market Cap of each supplier:", mkt_cap_per_supplier)
print("\n==> Products with low inventory:", low_inventory)
